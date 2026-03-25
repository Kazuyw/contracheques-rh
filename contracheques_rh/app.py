import os
import re
import smtplib
from datetime import datetime
from email.message import EmailMessage

from flask import (
    Flask, render_template, request, redirect,
    url_for, session, flash
)
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from email_validator import validate_email, EmailNotValidError
from dotenv import load_dotenv

from config import Config

load_dotenv()

app = Flask(__name__)
app.config.from_object(Config)

# Guardar temporariamente o último processamento em memória
ULTIMO_RESULTADO = []


def garantir_pastas():
    os.makedirs(app.config["UPLOAD_PLANILHAS"], exist_ok=True)
    os.makedirs(app.config["UPLOAD_PDFS"], exist_ok=True)
    os.makedirs(os.path.dirname(app.config["LOG_FILE"]), exist_ok=True)


def logar(mensagem: str):
    with open(app.config["LOG_FILE"], "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now().strftime('%d/%m/%Y %H:%M:%S')}] {mensagem}\n")


def usuario_logado():
    return session.get("logado", False)


def validar_email_formato(email: str):
    try:
        validado = validate_email(email, check_deliverability=False)
        return True, validado.normalized
    except EmailNotValidError as e:
        return False, str(e)


def ler_planilha(caminho_planilha):
    wb = load_workbook(caminho_planilha)
    ws = wb.active

    dados = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        nome = str(row[0]).strip() if row[0] else ""
        email = str(row[1]).strip() if row[1] else ""
        pdf_nome = str(row[2]).strip() if row[2] else ""

        dados.append({
            "linha": i,
            "nome": nome,
            "email": email,
            "pdf_nome": pdf_nome
        })

    return dados


def validar_registros(registros):
    pdfs_enviados = set(os.listdir(app.config["UPLOAD_PDFS"]))
    resultado = []

    for item in registros:
        erros = []

        if not item["nome"]:
            erros.append("Nome vazio.")

        ok_email, email_info = validar_email_formato(item["email"])
        if not ok_email:
            erros.append(f"E-mail inválido: {email_info}")

        if not item["pdf_nome"]:
            erros.append("Nome do PDF vazio.")
        elif item["pdf_nome"] not in pdfs_enviados:
            erros.append("PDF não encontrado na pasta de uploads.")

        resultado.append({
            "linha": item["linha"],
            "nome": item["nome"],
            "email": item["email"],
            "pdf_nome": item["pdf_nome"],
            "status_validacao": "OK" if not erros else "ERRO",
            "erros": erros
        })

    return resultado


def enviar_email_com_anexo(destinatario, nome_funcionario, caminho_pdf):
    msg = EmailMessage()
    msg["Subject"] = "Seu contracheque"
    msg["From"] = app.config["SMTP_SENDER"]
    msg["To"] = destinatario

    corpo = f"""Olá, {nome_funcionario}.

Segue em anexo o seu contracheque.

Atenciosamente,
RH - Colégio Interativo
"""
    msg.set_content(corpo)

    with open(caminho_pdf, "rb") as f:
        dados_pdf = f.read()
        nome_arquivo = os.path.basename(caminho_pdf)

    msg.add_attachment(
        dados_pdf,
        maintype="application",
        subtype="pdf",
        filename=nome_arquivo
    )

    with smtplib.SMTP(app.config["SMTP_SERVER"], app.config["SMTP_PORT"]) as server:
        server.starttls()
        server.login(app.config["SMTP_USER"], app.config["SMTP_PASS"])
        server.send_message(msg)


@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        senha = request.form.get("senha", "")
        if senha == app.config["APP_PASSWORD"]:
            session["logado"] = True
            return redirect(url_for("dashboard"))
        flash("Senha inválida.", "erro")
    return render_template("login.html")


@app.route("/dashboard", methods=["GET", "POST"])
def dashboard():
    global ULTIMO_RESULTADO

    if not usuario_logado():
        return redirect(url_for("login"))

    if request.method == "POST":
        planilha = request.files.get("planilha")
        pdfs = request.files.getlist("pdfs")

        if not planilha:
            flash("Envie a planilha Excel.", "erro")
            return redirect(url_for("dashboard"))

        if not pdfs or pdfs == [None]:
            flash("Envie os PDFs dos contracheques.", "erro")
            return redirect(url_for("dashboard"))

        # Limpa uploads anteriores
        for pasta in [app.config["UPLOAD_PLANILHAS"], app.config["UPLOAD_PDFS"]]:
            for arquivo in os.listdir(pasta):
                os.remove(os.path.join(pasta, arquivo))

        # Salva planilha
        nome_planilha = secure_filename(planilha.filename)
        caminho_planilha = os.path.join(app.config["UPLOAD_PLANILHAS"], nome_planilha)
        planilha.save(caminho_planilha)

        # Salva PDFs
        for pdf in pdfs:
            if pdf and pdf.filename:
                nome_pdf = secure_filename(pdf.filename)
                caminho_pdf = os.path.join(app.config["UPLOAD_PDFS"], nome_pdf)
                pdf.save(caminho_pdf)

        registros = ler_planilha(caminho_planilha)
        ULTIMO_RESULTADO = validar_registros(registros)

        return render_template("resultado.html", resultados=ULTIMO_RESULTADO)

    return render_template("dashboard.html")


@app.route("/enviar", methods=["POST"])
def enviar():
    global ULTIMO_RESULTADO

    if not usuario_logado():
        return redirect(url_for("login"))

    sucessos = 0
    erros = 0

    for item in ULTIMO_RESULTADO:
        if item["status_validacao"] != "OK":
            logar(f"ERRO VALIDAÇÃO | Linha {item['linha']} | {item['nome']} | {item['email']} | {'; '.join(item['erros'])}")
            erros += 1
            continue

        try:
            caminho_pdf = os.path.join(app.config["UPLOAD_PDFS"], item["pdf_nome"])
            enviar_email_com_anexo(item["email"], item["nome"], caminho_pdf)
            logar(f"SUCESSO | {item['nome']} | {item['email']} | {item['pdf_nome']}")
            sucessos += 1
        except Exception as e:
            logar(f"ERRO ENVIO | {item['nome']} | {item['email']} | {item['pdf_nome']} | {str(e)}")
            erros += 1

    flash(f"Envio finalizado. Sucessos: {sucessos} | Erros: {erros}", "info")
    return redirect(url_for("dashboard"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


if __name__ == "__main__":
    garantir_pastas()
    app.run(host="0.0.0.0", port=5000, debug=True)